#!/usr/bin/python3.5

# sheet_to_text.py - Program opens a spreadsheet and writes the cells of column A into on text file, The cells
# of column B into another text file, and so on

# Will take a excel file as input

import openpyxl
import re
import os
import sys
from openpyxl.utils import get_column_letter

# get excel file from command line
excel_file = sys.argv[1]

# open workbook
wb = openpyxl.load_workbook(excel_file)
sheet = wb.get_active_sheet()

# loop through excel file. opens a text file to write the column to.abs
for col_num in range(1, sheet.max_column + 1):
    col_letter = get_column_letter(col_num)
    text_file = open('colum_' + col_letter, 'w')
    for row_num in range(1, sheet.max_row + 1):
        text_file.write(str(sheet[col_letter + str(row_num)].value))
    text_file.close()