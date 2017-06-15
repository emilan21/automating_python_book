#!/usr/bin/env python
# -*- coding: utf-8 -*-

# update_produce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their update prices
PRICE_UPDATES = {'Garlic': 7.07,
                 'Celery': 2.19,
                 'Lemon': 1.27}

# Loop throught the rows and update the prices.
for row_num in range(2, sheet.max_row):         # skip first row
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updated_produce_sales.xlsx')
