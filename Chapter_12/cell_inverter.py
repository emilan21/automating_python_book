#!/usr/bin/env python
# -*- coding: utf-8 -*-

# cell_inverter.py - Inverts the row and column of the cells in a spread sheet

import openpyxl
from openpyxl.utils import get_column_letter
import sys

work_book = sys.argv[1]

original_wb = openpyxl.load_workbook(work_book)
original_sheet = original_wb.get_active_sheet()

new_wb = openpyxl.Workbook()
new_sheet = new_wb.get_active_sheet()

sheet_data = [[]]

# copy cell values into list
for row_num in range(1, original_sheet.max_row + 1):
    sheet_data.append(row_num)
    for col_num in range(1, original_sheet.max_column + 1):
        sheet_data[row_num].append(original_sheet.cell(row=row_num, column=col_num).value)     
        print(original_sheet.cell(row=row_num, column=col_num).value)        

print(sheet_data)
