#!/usr/bin/env python
# -*- coding: utf-8 -*-

#blank_row_inserter.py - Takes two integers and a filename string from command line. First integer
# is N and the second integer is M. Starting at row N, the program will insert M blank rows into the
# spreadsheet.

import openpyxl
from openpyxl.utils import get_column_letter
import sys

srow = int(sys.argv[1])
insert_num = int(sys.argv[2])
work_book = sys.argv[3]

original_wb = openpyxl.load_workbook(work_book)
original_sheet = original_wb.get_active_sheet()

new_wb = openpyxl.Workbook()
new_sheet = new_wb.get_active_sheet()

# copy up to N
for row_num in range(1, srow):
    for col_num in range(1, original_sheet.max_column + 1):
        new_sheet.cell(row=row_num, column=col_num).value = original_sheet[get_column_letter(col_num) + str(row_num)].value

for after_row in range(srow + insert_num, original_sheet.max_row + 1):
    for after_col in range(1, original_sheet.max_column + 1):
        new_sheet.cell(row=after_row, column=after_col).value = original_sheet[get_column_letter(after_col) + str(after_row)].value

new_wb.save('new_' + work_book)
