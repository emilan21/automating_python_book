#!/usr/bin/env python
# -*- coding: utf-8 -*-

# multiplication_table.py - Takes a number from the command line adn creates an NxN
# multiplication table in an Excel spreadhseet.

import openpyxl
from openpyxl.utils import get_column_letter
import sys

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

input_number = sys.argv[1]
count = 1

# Format top row
for column_num in range(2, int(input_number) + 2):
    if count < int(input_number) + 1:
        sheet[get_column_letter(column_num) + '1'] = count
        count += 1

count = 1
# Format first column
for row_num in range(2, int(input_number) + 2):
    if count < int(input_number) + 1:
        sheet['A' + str(count + 1)] = count
        count += 1

# Make multiplication table
for row_num in range(2, int(input_number) + 2):
    for column_num in range(2, int(input_number) + 2):
        sheet.cell(row=row_num, column=column_num).value = ((row_num - 1) * (column_num - 1))

# Save workbook
wb.save('multiplication_table.xlsx')
