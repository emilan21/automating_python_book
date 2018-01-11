#!/usr/bin/python3.5

# excel_to_csv.py - Converts from excel to csv in current directory. One csv file per file per sheet. 
# Will name the file "<excelfilename>_<sheettitle>.csv"

import csv
import os
import openpyxl

os.makedirs('converted_excel_files', exist_ok=True)

for excel_file in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if excel_file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excel_file)

        for sheet_name in wb.get_sheet_names():
            # Loop through every sheet in the workbook.
            sheet = wb.get_sheet_by_name(sheet_name)

            # Create the CSV filename from the Excel filename and sheet title.
            csv_file_name = os.path.splitext(excel_file)[0] + '_' + sheet_name + '.csv'

            # Create the csv.writer object for this CSV file.
            csv_file_obj = open(os.path.join('converted_excel_files', csv_file_name), 'w', newline='')
            csv_writer = csv.writer(csv_file_obj)

            # Loop through every row in the sheet.
            for row_num in range(1, sheet.max_row + 1):
                row_data = []       # append each cell to this list
                # Loop through each cell in the row.
                for col_num in range(1, sheet.max_column + 1):
                    row_data.append(sheet.cell(row=row_num, column=col_num).value) 
                # Write the row_data list to the csv file.
                csv_writer.writerow(row_data)
            csv_file_obj.close()